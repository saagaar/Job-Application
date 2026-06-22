from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from db.models import Job

XLSX_PATH = Path(__file__).parent.parent / "jobs" / "jobs_tracker.xlsx"

COLUMNS = [
    ("id",                  "ID"),
    ("title",               "Job Title"),
    ("company",             "Company"),
    ("match_score",         "Match %"),
    ("match_skills",        "Matching Skills"),
    ("match_gaps",          "Gaps"),
    ("match_reasoning",     "Reasoning"),
    ("status",              "Status"),
    ("location",            "Location"),
    ("salary_range",        "Salary"),
    ("date_found",          "Date Found"),
    ("applied_date",        "Applied"),
    ("source",              "Source"),
    ("url",                 "URL"),
    ("notes",               "Notes"),
    ("cv_path",                 "CV"),
    ("cover_letter_path",       "Cover Letter PDF"),
    ("cover_letter_content",    "Cover Letter"),
    ("description",             "Job Content"),
]

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
STATUS_VALUES = '"new,applied,interview,offer,rejected"'


class ExcelExporter:
    def export(self, jobs: list[Job], output_path: Path = XLSX_PATH) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jobs"

        # Header row
        for col_idx, (field, label) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 20

        # Data rows
        score_col = next(i for i, (f, _) in enumerate(COLUMNS, 1) if f == "match_score")
        status_col = next(i for i, (f, _) in enumerate(COLUMNS, 1) if f == "status")
        url_col = next(i for i, (f, _) in enumerate(COLUMNS, 1) if f == "url")
        desc_col = next(i for i, (f, _) in enumerate(COLUMNS, 1) if f == "description")
        cl_col = next(i for i, (f, _) in enumerate(COLUMNS, 1) if f == "cover_letter_content")

        for row_idx, job in enumerate(jobs, start=2):
            for col_idx, (field, _) in enumerate(COLUMNS, start=1):
                val = getattr(job, field, None)
                if field in ("match_skills", "match_gaps") and val:
                    try:
                        val = ", ".join(json.loads(val))
                    except Exception:
                        pass
                if hasattr(val, "value"):
                    val = val.value
                if hasattr(val, "isoformat"):
                    val = val.strftime("%Y-%m-%d")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if col_idx == url_col and val:
                    cell.hyperlink = val
                    cell.font = Font(color="0563C1", underline="single")
                if col_idx in (desc_col, cl_col):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    ws.row_dimensions[row_idx].height = 80

        num_rows = len(jobs)
        if num_rows > 0:
            self._apply_score_formatting(ws, score_col, num_rows)
            self._add_status_dropdown(ws, status_col, num_rows)

        # Auto-width columns (description gets a fixed width — it's too long to auto-size)
        for col_idx, (field, label) in enumerate(COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            if field in ("description", "cover_letter_content"):
                ws.column_dimensions[col_letter].width = 80
                continue
            max_len = len(label)
            for row_idx in range(2, num_rows + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, min(len(str(val)), 50))
            ws.column_dimensions[col_letter].width = max_len + 2

        wb.save(output_path)
        return output_path

    def _apply_score_formatting(self, ws, score_col: int, num_rows: int) -> None:
        col_letter = get_column_letter(score_col)
        score_range = f"{col_letter}2:{col_letter}{num_rows + 1}"
        rule = ColorScaleRule(
            start_type="num", start_value=0, start_color="FF0000",
            mid_type="num", mid_value=60, mid_color="FFFF00",
            end_type="num", end_value=100, end_color="00B050",
        )
        ws.conditional_formatting.add(score_range, rule)

    def _add_status_dropdown(self, ws, status_col: int, num_rows: int) -> None:
        col_letter = get_column_letter(status_col)
        dv = DataValidation(type="list", formula1=STATUS_VALUES, allow_blank=True)
        dv.sqref = f"{col_letter}2:{col_letter}{num_rows + 1}"
        ws.add_data_validation(dv)
